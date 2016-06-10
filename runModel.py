#This program will run a saved model on the given corpus and return which topics are relevant for each paper.
#Copyright (C) 2016 Magnus Johansson <magnusjohansson82(at@)gmail(dot.)com> (Enekullegatan 12, 418 75 Göteborg), Jonathan Klemetz <jonathanklemetz(at@)gmail(dot.)com> (Färgfabriksgatan 18, 417 24 Göteborg)
#This program is free software; you can redistribute it and/or modify it under the terms of the GNU General Public License as published by the Free Software Foundation; either version 2 of the License, or (at your option) any later version.
#This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU General Public License for more details.
#You should have received a copy of the GNU General Public License along with this program; if not, write to the Free Software Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA 02111-1307 USA


import gensim, os, bz2, inspect, logging, sys
from gensim import corpora, models, similarities
from nltk import word_tokenize
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Most popular topics"
number_of_topics = 3

def store_papername(name, row_number):
    ws.cell(column=1, row=row_number, value=name)

def store_probable(data, row_number):
    col = 2
    for i in range(0,number_of_topics):
        if(len(data) != 0):
            ws.cell(column = col+i, row=row_number, value=str(data[i]))
        else:
            pass

def main():
    top_directory = '/home/jonathan/Documents/Stuff/FixedLemmatizedHang'
    logging.basicConfig(format='%(asctime)s : %(levelname)s : %(message)s', level=logging.INFO)

    lda = gensim.models.LdaModel.load('/home/jonathan/Documents/Stuff/Models/hangsomething.lda')
    lda.print_topics(num_topics=100, num_words=7)
    row_number = 1
    for root, dirs, files in os.walk(top_directory):
        for file in filter(lambda file: file.endswith('.txt'), files):
            store_papername(file, row_number)
            print(file)
            originalString = open(os.path.join(root, file)).read()
            document = originalString
            document = word_tokenize(document)
            bow = lda.id2word.doc2bow(document)
            topic_analysis = lda[bow]
            sorted_analysis = sorted(topic_analysis, key = lambda x: x[1], reverse=True)
            print(sorted_analysis)
            store_probable(sorted_analysis, row_number)
            row_number = row_number+1
            print(topic_analysis)

main()
name_of_workbook = 'most_popular_topics.xlsx'
wb.save(filename = name_of_workbook)
